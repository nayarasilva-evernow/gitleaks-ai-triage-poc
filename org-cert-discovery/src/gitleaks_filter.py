from __future__ import annotations

from pathlib import Path
from typing import Iterable, List

from openpyxl import load_workbook

from progress import ProgressLog

DETAIL_SHEETS = ("DETALHES", "ACHADOS_CRITICOS", "POSSIVEIS_FP")


def filter_gitleaks_excel(
    path: Path,
    *,
    keywords: Iterable[str],
    progress: ProgressLog,
) -> List[dict]:
    if not path.exists():
        progress.warning(f"Arquivo Gitleaks não encontrado: {path}")
        return []

    progress.info(f"Abrindo foto Gitleaks: {path.name}")
    wb = load_workbook(path, read_only=True, data_only=True)
    keywords_l = [k.lower() for k in keywords]
    rows_out: list[dict] = []

    sheet_names = [s for s in DETAIL_SHEETS if s in wb.sheetnames]
    if not sheet_names and "DETALHES" not in wb.sheetnames:
        # fallback: primeira aba com Tipo_Segredo
        for name in wb.sheetnames:
            ws = wb[name]
            header = next(ws.iter_rows(values_only=True), None)
            if header and "Tipo_Segredo" in header:
                sheet_names = [name]
                break

    if not sheet_names:
        progress.warning("Nenhuma aba com Tipo_Segredo encontrada no Excel.")
        wb.close()
        return []

    # Prefer DETALHES se existir (evita duplicar críticas + detalhes)
    if "DETALHES" in sheet_names:
        sheet_names = ["DETALHES"]

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            continue
        cols = [str(c) if c is not None else "" for c in header]
        try:
            idx_tipo = cols.index("Tipo_Segredo")
        except ValueError:
            progress.warning(f"Aba {sheet_name} sem coluna Tipo_Segredo.")
            continue

        col_index = {c: i for i, c in enumerate(cols) if c}
        total_seen = 0
        matched = 0
        for row in it:
            if not row or row[0] is None:
                continue
            total_seen += 1
            tipo = str(row[idx_tipo] or "")
            tipo_l = tipo.lower()
            if not any(k in tipo_l for k in keywords_l):
                continue
            matched += 1
            item = {
                "fonte": f"Gitleaks:{sheet_name}",
                "repo": _cell(row, col_index, "Repo"),
                "linguagem": _cell(row, col_index, "Linguagem"),
                "topics": _cell(row, col_index, "Topics"),
                "arquivo": _cell(row, col_index, "Arquivo"),
                "linha": _cell(row, col_index, "Linha"),
                "tipo_segredo": tipo,
                "classificacao": _cell(row, col_index, "Classificacao"),
                "motivo": _cell(row, col_index, "Motivo"),
                "descricao": _cell(row, col_index, "Descricao"),
                # Preview omitido do relatório executivo por segurança;
                # mantido truncado na planilha detalhada se necessário
                "preview": _truncate(_cell(row, col_index, "Preview"), 80),
                "fingerprint": _cell(row, col_index, "Fingerprint"),
                "categoria_demanda": _categoria(tipo_l),
                "autoassinado": "N/D",
                "wildcard": "N/D",
                "acao_sugerida": _acao_gitleaks(tipo_l),
            }
            rows_out.append(item)
            if matched % 500 == 0:
                progress.progress(matched, 0, label=f"matches em {sheet_name}")

        progress.info(
            f"Aba {sheet_name}: {total_seen} linhas lidas | {matched} cert/chave."
        )

    wb.close()
    return rows_out


def _cell(row, col_index: dict, name: str) -> str:
    i = col_index.get(name)
    if i is None or i >= len(row):
        return ""
    v = row[i]
    return "" if v is None else str(v)


def _truncate(s: str, n: int) -> str:
    s = s or ""
    return s if len(s) <= n else s[: n - 3] + "..."


def _categoria(tipo_l: str) -> str:
    if "private" in tipo_l or "rsa" in tipo_l or "ssh" in tipo_l:
        return "chave_privada"
    if "cert" in tipo_l or "pem" in tipo_l:
        return "certificado"
    if "key" in tipo_l:
        return "chave_ou_segredo"
    return "outros_relacionados"


def _acao_gitleaks(tipo_l: str) -> str:
    if "private" in tipo_l:
        return "Remover do repositório e rotacionar a chave; tratar como incidente."
    return "Validar se é material sensível; remover do VCS se aplicável."
