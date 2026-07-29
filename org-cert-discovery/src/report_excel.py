from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from progress import ProgressLog

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SEVERITY_FILL = {
    "Crítica": PatternFill("solid", fgColor="FF6B6B"),
    "Alta": PatternFill("solid", fgColor="FFB347"),
    "Média": PatternFill("solid", fgColor="FFE066"),
    "Informativa": PatternFill("solid", fgColor="C0C0C0"),
}


def write_excel(
    path: Path,
    *,
    inventory_rows: List[dict],
    gitleaks_rows: List[dict],
    org: str,
    progress: ProgressLog,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    _write_sumario(wb.active, inventory_rows, gitleaks_rows, org)
    _write_jks(wb.create_sheet("JKS_INVENTARIO"), inventory_rows)
    _write_gitleaks(wb.create_sheet("CERTIFICADOS_CHAVES"), gitleaks_rows)
    _write_por_repo(wb.create_sheet("Visao_por_Repositorio"), inventory_rows, gitleaks_rows)
    _write_acoes(wb.create_sheet("Acoes_Sugeridas"), inventory_rows, gitleaks_rows)
    _write_legenda(wb.create_sheet("Legenda_Metodo"))

    wb.save(path)
    progress.info(f"Excel gerado: {path}")
    return path


def _style_header(ws, headers: list[str]) -> None:
    for col, name in enumerate(headers, 1):
        cell = ws.cell(1, col, name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col[:80]:
            length = max(length, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(10, length + 2), max_width)


def _severity_for_artifact(artifact_type: str) -> str:
    if artifact_type in {"JKS", "Keystore", "P12/PFX"}:
        return "Alta"
    if artifact_type == "Cert/Chave (arquivo)":
        return "Média"
    return "Informativa"


def _write_sumario(ws, inventory_rows, gitleaks_rows, org: str) -> None:
    ws.title = "Sumario_Executivo"
    jks = [r for r in inventory_rows if r.get("artifact_type") in {"JKS", "Keystore", "P12/PFX"}]
    cert_files = [r for r in inventory_rows if r.get("artifact_type") == "Cert/Chave (arquivo)"]
    repos_inv = {r.get("repo") for r in inventory_rows}
    repos_gl = {r.get("repo") for r in gitleaks_rows}

    metrics = [
        ("Data_geracao_UTC", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Organizacao", org),
        ("Permissao_token", "Somente leitura (GET) — nenhuma alteração nos repositórios"),
        ("Total_arquivos_inventario_API", len(inventory_rows)),
        ("Total_JKS_Keystore_P12", len(jks)),
        ("Total_cert_chave_arquivo", len(cert_files)),
        ("Repos_com_artefato_inventario", len(repos_inv)),
        ("Total_gitleaks_cert_chave", len(gitleaks_rows)),
        ("Repos_gitleaks_cert_chave", len(repos_gl)),
        (
            "Autoassinado_Wildcard_no_JKS",
            "N/D nesta entrega — requer abertura do keystore (fase 2)",
        ),
        (
            "Escopo_desta_entrega",
            "Inventário path + accountability (último commit) + fatia Gitleaks local",
        ),
        (
            "Proxima_fase",
            "Abertura JKS com rol de senhas (ferramenta dedicada) + parse PEM para autoassinado/wildcard",
        ),
    ]
    _style_header(ws, ["Metrica", "Valor"])
    for i, (k, v) in enumerate(metrics, 2):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)
    _autosize(ws, 80)


def _write_jks(ws, inventory_rows) -> None:
    headers = [
        "ID",
        "Severidade_provisoria",
        "Tipo_artefato",
        "Repositorio",
        "Caminho_arquivo",
        "URL_arquivo",
        "Tamanho_bytes",
        "Ultimo_autor",
        "Ultimo_autor_email",
        "Data_ultimo_commit",
        "Hash_commit",
        "Mensagem_commit",
        "Autoassinado",
        "Wildcard",
        "Status_analise",
        "Acao_sugerida",
        "Observacao",
    ]
    _style_header(ws, headers)
    for i, r in enumerate(inventory_rows, 1):
        sev = _severity_for_artifact(r.get("artifact_type") or "")
        is_ks = r.get("artifact_type") in {"JKS", "Keystore", "P12/PFX"}
        row = [
            f"INV-{i:04d}",
            sev,
            r.get("artifact_type") or "",
            r.get("repo") or "",
            r.get("path") or "",
            r.get("html_url") or "",
            r.get("size") or "",
            r.get("author_name") or "",
            r.get("author_email") or "",
            r.get("commit_date") or "",
            r.get("commit_sha") or "",
            r.get("commit_message") or "",
            "N/D — fase 2",
            "N/D — fase 2",
            "Localizado (read-only)" if r.get("author_done") else "Localizado (autor pendente)",
            (
                "Remover keystore do VCS; rotacionar material; migrar para keystore corporativo."
                if is_ks
                else "Validar se há chave privada; remover do VCS se sensível."
            ),
            "Classificação autoassinado/wildcard depende de abertura/parse (não feito nesta corrida).",
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(i + 1, col, val)
            if col == 2 and sev in SEVERITY_FILL:
                cell.fill = SEVERITY_FILL[sev]
    _autosize(ws)


def _write_gitleaks(ws, gitleaks_rows) -> None:
    headers = [
        "ID",
        "Fonte",
        "Repositorio",
        "Arquivo",
        "Linha",
        "Tipo_Segredo",
        "Categoria_demanda",
        "Classificacao",
        "Motivo",
        "Descricao",
        "Fingerprint",
        "Preview_truncado",
        "Autoassinado",
        "Wildcard",
        "Acao_sugerida",
    ]
    _style_header(ws, headers)
    for i, r in enumerate(gitleaks_rows, 1):
        row = [
            f"GL-{i:04d}",
            r.get("fonte") or "",
            r.get("repo") or "",
            r.get("arquivo") or "",
            r.get("linha") or "",
            r.get("tipo_segredo") or "",
            r.get("categoria_demanda") or "",
            r.get("classificacao") or "",
            r.get("motivo") or "",
            r.get("descricao") or "",
            r.get("fingerprint") or "",
            r.get("preview") or "",
            r.get("autoassinado") or "N/D",
            r.get("wildcard") or "N/D",
            r.get("acao_sugerida") or "",
        ]
        for col, val in enumerate(row, 1):
            ws.cell(i + 1, col, val)
    _autosize(ws)


def _write_por_repo(ws, inventory_rows, gitleaks_rows) -> None:
    headers = [
        "Repositorio",
        "Qtd_inventario",
        "Qtd_JKS",
        "Qtd_cert_arquivo",
        "Qtd_gitleaks",
        "Autores_distintos",
        "Severidade_maxima",
        "Prioridade_followup",
    ]
    _style_header(ws, headers)

    inv_by: dict[str, list] = {}
    for r in inventory_rows:
        inv_by.setdefault(r.get("repo") or "", []).append(r)
    gl_by: dict[str, int] = Counter(r.get("repo") or "" for r in gitleaks_rows)

    repos = sorted(set(inv_by) | set(gl_by))
    severity_rank = {"Crítica": 4, "Alta": 3, "Média": 2, "Informativa": 1}

    for i, repo in enumerate(repos, 1):
        items = inv_by.get(repo, [])
        jks = sum(1 for x in items if x.get("artifact_type") in {"JKS", "Keystore", "P12/PFX"})
        certs = sum(1 for x in items if x.get("artifact_type") == "Cert/Chave (arquivo)")
        authors = {x.get("author_name") for x in items if x.get("author_name")}
        sev = "Informativa"
        for x in items:
            s = _severity_for_artifact(x.get("artifact_type") or "")
            if severity_rank[s] > severity_rank[sev]:
                sev = s
        if gl_by.get(repo, 0) and severity_rank.get("Crítica", 0) >= severity_rank[sev]:
            # presença de chave privada no gitleaks sobe prioridade textual
            pass
        gl_n = gl_by.get(repo, 0)
        if jks or gl_n:
            sev = "Alta" if severity_rank[sev] < 3 else sev
        priority = 1
        if jks >= 3 or gl_n >= 5:
            priority = 5
        elif jks or gl_n >= 2:
            priority = 4
        elif items:
            priority = 3
        row = [repo, len(items), jks, certs, gl_n, len(authors), sev, priority]
        for col, val in enumerate(row, 1):
            ws.cell(i + 1, col, val)
    _autosize(ws)


def _write_acoes(ws, inventory_rows, gitleaks_rows) -> None:
    headers = ["Prioridade", "Alvo", "Tipo", "Acao_sugerida", "Responsavel_indicado", "Status"]
    _style_header(ws, headers)
    lines = []
    for r in inventory_rows:
        if r.get("artifact_type") in {"JKS", "Keystore", "P12/PFX"}:
            lines.append(
                (
                    1,
                    f"{r.get('repo')}:{r.get('path')}",
                    r.get("artifact_type"),
                    "Remover do VCS; rotacionar; não reintroduzir keystore no código.",
                    r.get("author_name") or "Time do repositório",
                    "Aberto",
                )
            )
    for r in gitleaks_rows:
        if "private" in (r.get("tipo_segredo") or "").lower():
            lines.append(
                (
                    1,
                    f"{r.get('repo')}:{r.get('arquivo')}",
                    r.get("tipo_segredo"),
                    r.get("acao_sugerida"),
                    "Time do repositório",
                    "Aberto",
                )
            )
    lines = lines[:500]
    for i, row in enumerate(lines, 1):
        for col, val in enumerate(row, 1):
            ws.cell(i + 1, col, val)
    _autosize(ws)


def _write_legenda(ws) -> None:
    _style_header(ws, ["Item", "Descricao"])
    rows = [
        ("Token", "Somente leitura. Cliente bloqueia qualquer método além de GET/HEAD."),
        ("JKS_INVENTARIO", "Arquivos encontrados na árvore default do repositório via API."),
        ("Autoassinado/Wildcard", "Não classificados nesta corrida (exige abertura/parse)."),
        ("Ultimo_autor", "Autor do último commit que tocou o path (accountability inicial)."),
        ("CERTIFICADOS_CHAVES", "Fatia da foto Gitleaks filtrada por Tipo_Segredo."),
        ("Retomada", "Checkpoint SQLite em data/checkpoints — rerode o mesmo comando após queda."),
        ("Rate limit", "Script pausa automaticamente quando o budget da API está baixo."),
    ]
    for i, (a, b) in enumerate(rows, 2):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
    _autosize(ws, 90)
